"""
API endpoints for React frontend
Provides JSON responses for dashboard, employees, departments, and accounts
"""
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Count  # Keep for potential future use
from datetime import timedelta
import json

from ..models import Employee, AttendanceRecord, Department
from .utils import get_vietnam_now


def json_response(data, status=200):
    """Helper to create JSON response"""
    return JsonResponse(data, status=status, json_dumps_params={'ensure_ascii': False})


def error_response(message, status=400):
    """Helper for error responses"""
    return json_response({'success': False, 'message': message}, status)


@csrf_exempt
def dashboard_api(request):
    """API to get dashboard data for React frontend"""
    if request.method != 'GET':
        return error_response('Method not allowed', 405)
    
    from django.core.cache import cache
    
    # Try to get from cache first (cache stats for 30 seconds)
    cache_key = 'dashboard_stats_v1'
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return json_response(cached_data)
    
    vietnam_now = get_vietnam_now()
    
    # Company statistics
    company_stats = {
        'total_employees': Employee.objects.count(),
        'working': Employee.objects.filter(work_status='WORKING').count(),
        'on_leave': Employee.objects.filter(work_status='ON_LEAVE').count(),
        'terminated': Employee.objects.filter(work_status='TERMINATED').count(),
        'in_office': Employee.objects.filter(current_status='IN_OFFICE', work_status='WORKING').count(),
        'out_office': Employee.objects.filter(current_status='OUT_OFFICE', work_status='WORKING').count(),
        'not_in': Employee.objects.filter(current_status='NOT_IN', work_status='WORKING').count(),
    }
    
    # Recent employees with attendance
    yesterday = vietnam_now.date() - timedelta(days=1)
    recent_records = AttendanceRecord.objects.filter(
        date__gte=yesterday
    ).select_related('employee', 'employee__user').order_by('-check_in_time')
    
    recent_employee_ids = list(recent_records.values_list('employee_id', flat=True).distinct()[:10])
    employees_qs = Employee.objects.filter(id__in=recent_employee_ids).select_related('user')
    
    employees = []
    for emp in employees_qs:
        employees.append({
            'employee_id': emp.employee_id,
            'full_name': emp.get_full_name(),
            'department': emp.department,
            'position': emp.position,
            'work_status': emp.work_status,
            'current_status': emp.current_status,
            'has_face': emp.face_embeddings is not None,
        })
    
    response_data = {
        'success': True,
        'company_stats': company_stats,
        'employees': employees,
        'total_departments': Department.objects.count(),
        'current_time': vietnam_now.isoformat(),
    }
    
    # Cache for 30 seconds
    cache.set(cache_key, response_data, 30)
    
    return json_response(response_data)


@csrf_exempt
def employees_api(request):
    """API to list/create employees"""
    if request.method == 'GET':
        employees = Employee.objects.select_related('user').all()
        
        # Apply filters
        department = request.GET.get('department')
        status = request.GET.get('status')
        search = request.GET.get('search', '')
        
        if department:
            employees = employees.filter(department=department)
        if status:
            employees = employees.filter(work_status=status)
        if search:
            employees = employees.filter(
                user__first_name__icontains=search
            ) | employees.filter(
                user__last_name__icontains=search
            ) | employees.filter(
                employee_id__icontains=search
            )
        
        data = []
        for emp in employees:
            data.append({
                'employee_id': emp.employee_id,
                'full_name': emp.get_full_name(),
                'email': emp.email or '',
                'department': emp.department,
                'position': emp.position,
                'work_status': emp.work_status,
                'work_status_display': emp.get_work_status_display(),
                'current_status': emp.current_status,
                'join_date': emp.join_date.isoformat() if emp.join_date else None,
                'has_face': emp.face_embeddings is not None,
                'has_account': emp.user is not None,
            })
        
        # Get departments for filter
        departments = list(Department.objects.values_list('name', flat=True))
        
        # Statistics
        stats = {
            'total': Employee.objects.count(),
            'working': Employee.objects.filter(work_status='WORKING').count(),
            'on_leave': Employee.objects.filter(work_status='ON_LEAVE').count(),
            'terminated': Employee.objects.filter(work_status='TERMINATED').count(),
        }
        
        return json_response({
            'success': True,
            'employees': data,
            'departments': departments,
            'stats': stats,
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Auto-generate employee_id if not provided
            employee_id = data.get('employee_id')
            if not employee_id:
                # Find the highest employee_id and increment
                last_employee = Employee.objects.filter(
                    employee_id__startswith='NV'
                ).order_by('-employee_id').first()
                
                if last_employee:
                    try:
                        last_num = int(last_employee.employee_id[2:])  # Extract number after 'NV'
                        new_num = last_num + 1
                    except ValueError:
                        new_num = 1
                else:
                    new_num = 1
                
                employee_id = f'NV{new_num:03d}'  # Format: NV001, NV002, etc.
            
            # Create employee without user account
            # User/account can be created separately via account management
            employee = Employee.objects.create(
                user=None,
                employee_id=employee_id,
                first_name=data.get('first_name', ''),
                last_name=data.get('last_name', ''),
                email=data.get('email', ''),
                department=data.get('department', ''),
                position=data.get('position', ''),
                work_status=data.get('work_status', 'WORKING'),
            )
            
            return json_response({
                'success': True,
                'message': f'Đã tạo nhân viên {employee_id} thành công',
                'employee_id': employee.employee_id,
            }, 201)
            
        except json.JSONDecodeError:
            return error_response('Invalid JSON')
        except Exception as e:
            return error_response(str(e))
    
    return error_response('Method not allowed', 405)


@csrf_exempt
def employee_detail_api(request, employee_id):
    """API to get/update/delete single employee"""
    try:
        employee = Employee.objects.select_related('user').get(employee_id=employee_id)
    except Employee.DoesNotExist:
        return error_response('Employee not found', 404)
    
    if request.method == 'GET':
        # Get attendance history
        records = AttendanceRecord.objects.filter(
            employee=employee
        ).order_by('-date')[:30]
        
        history = []
        vietnam_tz = get_vietnam_now().tzinfo  # Get Vietnam timezone
        for record in records:
            # Convert UTC times to Vietnam timezone
            check_in_vn = None
            check_out_vn = None
            if record.check_in_time:
                check_in_vn = record.check_in_time.astimezone(vietnam_tz).strftime('%H:%M')
            if record.check_out_time:
                check_out_vn = record.check_out_time.astimezone(vietnam_tz).strftime('%H:%M')
            
            history.append({
                'date': record.date.isoformat(),
                'check_in': check_in_vn,
                'check_out': check_out_vn,
                'status': record.status,
                'status_display': record.get_status_display(),
            })
        
        return json_response({
            'success': True,
            'employee': {
                'employee_id': employee.employee_id,
                'username': employee.user.username if employee.user else None,
                'full_name': employee.get_full_name(),
                'first_name': employee.first_name,
                'last_name': employee.last_name,
                'email': employee.email or '',
                'department': employee.department,
                'position': employee.position,
                'work_status': employee.work_status,
                'work_status_display': employee.get_work_status_display(),
                'current_status': employee.current_status,
                'join_date': employee.join_date.isoformat() if employee.join_date else None,
                'has_face': employee.face_embeddings is not None,
                'has_account': employee.user is not None,
            },
            'attendance_history': history,
        })
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            
            # Update user (only if user exists)
            user = employee.user
            if user:
                if 'first_name' in data:
                    user.first_name = data['first_name']
                if 'last_name' in data:
                    user.last_name = data['last_name']
                if 'email' in data:
                    user.email = data['email']
                if 'password' in data and data['password']:
                    user.set_password(data['password'])
                user.save()
            
            # Update employee
            if 'department' in data:
                employee.department = data['department']
            if 'position' in data:
                employee.position = data['position']
            if 'work_status' in data:
                employee.work_status = data['work_status']
            employee.save()
            
            return json_response({
                'success': True,
                'message': 'Employee updated successfully',
            })
            
        except json.JSONDecodeError:
            return error_response('Invalid JSON')
        except Exception as e:
            return error_response(str(e))
    
    elif request.method == 'DELETE':
        user = employee.user
        employee.delete()
        if user:
            user.delete()
        return json_response({
            'success': True,
            'message': 'Đã xóa nhân viên thành công',
        })
    
    return error_response('Method not allowed', 405)


@csrf_exempt
def departments_api(request):
    """API to list/create departments"""
    if request.method == 'GET':
        departments = Department.objects.all()
        
        data = []
        for dept in departments:
            # Count employees by department name (since Employee.department is CharField)
            employee_count = Employee.objects.filter(department=dept.name).count()
            data.append({
                'id': dept.id,
                'name': dept.name,
                'description': dept.description,
                'employee_count': employee_count,
            })
        
        return json_response({
            'success': True,
            'departments': data,
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            dept = Department.objects.create(
                name=data['name'],
                description=data.get('description', ''),
            )
            
            return json_response({
                'success': True,
                'message': 'Department created successfully',
                'id': dept.id,
            }, 201)
            
        except json.JSONDecodeError:
            return error_response('Invalid JSON')
        except Exception as e:
            return error_response(str(e))
    
    return error_response('Method not allowed', 405)


@csrf_exempt
def department_detail_api(request, pk):
    """API to get/update/delete single department"""
    try:
        department = Department.objects.get(pk=pk)
    except Department.DoesNotExist:
        return error_response('Department not found', 404)
    
    # Count employees by department name
    employee_count = Employee.objects.filter(department=department.name).count()
    
    if request.method == 'GET':
        # Get employees in this department
        employees = Employee.objects.filter(
            department=department.name
        ).select_related('user')
        
        emp_data = []
        for emp in employees:
            emp_data.append({
                'employee_id': emp.employee_id,
                'full_name': emp.get_full_name(),
                'position': emp.position,
                'work_status': emp.work_status,
            })
        
        return json_response({
            'success': True,
            'department': {
                'id': department.id,
                'name': department.name,
                'description': department.description,
                'employee_count': employee_count,
            },
            'employees': emp_data,
        })
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            
            if 'name' in data:
                department.name = data['name']
            if 'description' in data:
                department.description = data['description']
            department.save()
            
            return json_response({
                'success': True,
                'message': 'Department updated successfully',
            })
            
        except json.JSONDecodeError:
            return error_response('Invalid JSON')
        except Exception as e:
            return error_response(str(e))
    
    elif request.method == 'DELETE':
        # Check if department has employees
        if employee_count > 0:
            return error_response(
                f'Không thể xóa phòng ban "{department.name}" vì còn {employee_count} nhân viên. '
                f'Vui lòng chuyển nhân viên sang phòng ban khác trước khi xóa.',
                400
            )
        
        department.delete()
        return json_response({
            'success': True,
            'message': 'Đã xóa phòng ban thành công',
        })
    
    return error_response('Method not allowed', 405)


@csrf_exempt
def accounts_api(request):
    """API to list/create user accounts"""
    if request.method == 'GET':
        users = User.objects.all().prefetch_related('employee')
        
        data = []
        for user in users:
            employee = getattr(user, 'employee', None)
            data.append({
                'id': user.id,
                'username': user.username,
                'full_name': user.get_full_name(),
                'email': user.email,
                'is_staff': user.is_staff,
                'is_superuser': user.is_superuser,
                'is_active': user.is_active,
                'date_joined': user.date_joined.isoformat(),
                'last_login': user.last_login.isoformat() if user.last_login else None,
                'has_employee': employee is not None,
                'employee_id': employee.employee_id if employee else None,
            })
        
        # Get employees without user accounts for account creation
        available_employees = Employee.objects.filter(user__isnull=True)
        available_list = []
        for emp in available_employees:
            available_list.append({
                'employee_id': emp.employee_id,
                'full_name': emp.get_full_name(),
                'first_name': emp.first_name,
                'last_name': emp.last_name,
                'email': emp.email or '',
                'department': emp.department,
                'position': emp.position,
            })
        
        return json_response({
            'success': True,
            'accounts': data,
            'available_employees': available_list,
        })
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Require employee_id for new accounts
            employee_id = data.get('employee_id')
            if not employee_id:
                return error_response('Vui lòng chọn nhân viên để liên kết tài khoản')
            
            try:
                employee = Employee.objects.get(employee_id=employee_id, user__isnull=True)
            except Employee.DoesNotExist:
                return error_response('Nhân viên không tồn tại hoặc đã có tài khoản')
            
            # Create user with employee's name
            user = User.objects.create_user(
                username=data['username'],
                password=data['password'],
                email=employee.email or '',
                first_name=employee.first_name,
                last_name=employee.last_name,
            )
            
            if data.get('is_staff'):
                user.is_staff = True
            if data.get('is_superuser'):
                user.is_superuser = True
            user.save()
            
            # Link employee to user
            employee.user = user
            employee.save()
            
            return json_response({
                'success': True,
                'message': f'Đã tạo tài khoản và liên kết với nhân viên {employee_id}',
                'id': user.id,
            }, 201)
            
        except json.JSONDecodeError:
            return error_response('Invalid JSON')
        except Exception as e:
            return error_response(str(e))
    
    return error_response('Method not allowed', 405)


@csrf_exempt
def account_detail_api(request, pk):
    """API to get/update/delete single account"""
    try:
        user = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return error_response('Account not found', 404)
    
    if request.method == 'GET':
        employee = getattr(user, 'employee', None)
        
        return json_response({
            'success': True,
            'account': {
                'id': user.id,
                'username': user.username,
                'full_name': user.get_full_name(),
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'is_staff': user.is_staff,
                'is_superuser': user.is_superuser,
                'is_active': user.is_active,
                'date_joined': user.date_joined.isoformat(),
                'last_login': user.last_login.isoformat() if user.last_login else None,
                'has_employee': employee is not None,
                'employee_id': employee.employee_id if employee else None,
            },
        })
    
    elif request.method == 'PUT':
        try:
            data = json.loads(request.body)
            
            # Cho phép đổi username
            if 'username' in data and data['username']:
                new_username = data['username']
                # Kiểm tra username mới không trùng với user khác
                if User.objects.filter(username=new_username).exclude(pk=user.pk).exists():
                    return error_response(f'Username "{new_username}" đã được sử dụng')
                user.username = new_username
            
            if 'is_staff' in data:
                user.is_staff = data['is_staff']
            if 'is_active' in data:
                user.is_active = data['is_active']
            if 'password' in data and data['password']:
                user.set_password(data['password'])
            user.save()
            
            return json_response({
                'success': True,
                'message': 'Account updated successfully',
            })
            
        except json.JSONDecodeError:
            return error_response('Invalid JSON')
        except Exception as e:
            return error_response(str(e))
    
    elif request.method == 'DELETE':
        # Unlink employee first if exists
        if hasattr(user, 'employee'):
            employee = user.employee
            employee.user = None
            employee.save()
        
        user.delete()
        return json_response({
            'success': True,
            'message': 'Đã xóa tài khoản thành công',
        })
    
    return error_response('Method not allowed', 405)


@csrf_exempt
def department_stats_api(request):
    """API to get department attendance statistics"""
    if request.method != 'GET':
        return error_response('Method not allowed', 405)
    
    from datetime import datetime
    from django.db.models import Count, Q
    
    vietnam_now = get_vietnam_now()
    
    # Get filter parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    department_filter = request.GET.get('department')
    
    # Default to current month
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = vietnam_now.date().replace(day=1)
    else:
        start_date = vietnam_now.date().replace(day=1)
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = vietnam_now.date()
    else:
        end_date = vietnam_now.date()
    
    # Get all departments
    departments = Department.objects.all()
    if department_filter:
        departments = departments.filter(name=department_filter)
    
    department_stats = []
    total_stats = {
        'total_employees': 0,
        'total_records': 0,
        'on_time': 0,
        'late': 0,
        'early': 0,
        'absent': 0,
    }
    
    for dept in departments:
        # Get employees in this department
        employees = Employee.objects.filter(department=dept.name, work_status='WORKING')
        employee_count = employees.count()
        
        # Get attendance records for this department within date range
        records = AttendanceRecord.objects.filter(
            employee__department=dept.name,
            date__gte=start_date,
            date__lte=end_date
        )
        
        record_count = records.count()
        on_time = records.filter(status='ON_TIME').count()
        late = records.filter(status='LATE').count()
        early = records.filter(status='EARLY').count()
        absent = records.filter(status='ABSENT').count()
        
        # Calculate attendance rate
        working_days = (end_date - start_date).days + 1
        expected_records = employee_count * working_days
        attendance_rate = (record_count / expected_records * 100) if expected_records > 0 else 0
        on_time_rate = (on_time / record_count * 100) if record_count > 0 else 0
        
        dept_data = {
            'id': dept.id,
            'name': dept.name,
            'employee_count': employee_count,
            'total_records': record_count,
            'on_time': on_time,
            'late': late,
            'early': early,
            'absent': absent,
            'attendance_rate': round(attendance_rate, 1),
            'on_time_rate': round(on_time_rate, 1),
        }
        department_stats.append(dept_data)
        
        # Update totals
        total_stats['total_employees'] += employee_count
        total_stats['total_records'] += record_count
        total_stats['on_time'] += on_time
        total_stats['late'] += late
        total_stats['early'] += early
        total_stats['absent'] += absent
    
    # Sort by employee count descending
    department_stats.sort(key=lambda x: x['employee_count'], reverse=True)
    
    return json_response({
        'success': True,
        'department_stats': department_stats,
        'total_stats': total_stats,
        'filters': {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'department': department_filter,
        },
        'departments': list(Department.objects.values_list('name', flat=True)),
    })


@csrf_exempt
def export_department_stats_excel(request):
    """Export department statistics to Excel file with detailed employee data"""
    if request.method != 'GET':
        return error_response('Method not allowed', 405)
    
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from datetime import datetime
    import io
    
    vietnam_now = get_vietnam_now()
    vietnam_tz = vietnam_now.tzinfo
    
    # Get filter parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    department_filter = request.GET.get('department')
    
    # Default to current month
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = vietnam_now.date().replace(day=1)
    else:
        start_date = vietnam_now.date().replace(day=1)
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = vietnam_now.date()
    else:
        end_date = vietnam_now.date()
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    dept_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Status colors
    status_fills = {
        'ON_TIME': PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        'LATE': PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        'EARLY': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    }
    
    # ===== SHEET 1: Summary by Department =====
    ws1 = wb.active
    ws1.title = "Tổng hợp phòng ban"
    
    # Title
    ws1.merge_cells('A1:H1')
    ws1['A1'] = "BÁO CÁO THỐNG KÊ CHẤM CÔNG THEO PHÒNG BAN"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = Alignment(horizontal="center")
    
    ws1.merge_cells('A2:H2')
    ws1['A2'] = f"Từ ngày {start_date.strftime('%d/%m/%Y')} đến ngày {end_date.strftime('%d/%m/%Y')}"
    ws1['A2'].alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['STT', 'Phòng ban', 'Số NV', 'Tổng lượt', 'Đúng giờ', 'Đi trễ', 'Về sớm', 'Tỷ lệ đúng giờ (%)']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Get departments
    departments = Department.objects.all()
    if department_filter:
        departments = departments.filter(name=department_filter)
    
    row = 5
    for idx, dept in enumerate(departments, 1):
        employees = Employee.objects.filter(department=dept.name, work_status='WORKING')
        employee_count = employees.count()
        
        records = AttendanceRecord.objects.filter(
            employee__department=dept.name,
            date__gte=start_date,
            date__lte=end_date
        )
        
        record_count = records.count()
        on_time = records.filter(status='ON_TIME').count()
        late = records.filter(status='LATE').count()
        early = records.filter(status='EARLY').count()
        on_time_rate = (on_time / record_count * 100) if record_count > 0 else 0
        
        data = [idx, dept.name, employee_count, record_count, on_time, late, early, round(on_time_rate, 1)]
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col in [3, 4, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center")
        row += 1
    
    # Adjust column widths for sheet 1
    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 10
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 18
    
    # ===== SHEET 2: Detailed Employee Attendance =====
    ws2 = wb.create_sheet(title="Chi tiết chấm công")
    
    # Title
    ws2.merge_cells('A1:G1')
    ws2['A1'] = "CHI TIẾT CHẤM CÔNG NHÂN VIÊN THEO PHÒNG BAN"
    ws2['A1'].font = Font(bold=True, size=16)
    ws2['A1'].alignment = Alignment(horizontal="center")
    
    ws2.merge_cells('A2:G2')
    ws2['A2'] = f"Từ ngày {start_date.strftime('%d/%m/%Y')} đến ngày {end_date.strftime('%d/%m/%Y')}"
    ws2['A2'].alignment = Alignment(horizontal="center")
    
    # Headers for detail sheet
    detail_headers = ['Phòng ban', 'Mã NV', 'Họ tên', 'Ngày', 'Giờ vào', 'Giờ ra', 'Trạng thái']
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    detail_row = 5
    
    for dept in departments:
        # Get employees in this department
        employees = Employee.objects.filter(
            department=dept.name,
            work_status='WORKING'
        ).select_related('user').order_by('employee_id')
        
        for emp in employees:
            # Get attendance records for this employee
            records = AttendanceRecord.objects.filter(
                employee=emp,
                date__gte=start_date,
                date__lte=end_date
            ).order_by('date')
            
            for record in records:
                # Convert times to Vietnam timezone
                check_in_str = ''
                check_out_str = ''
                if record.check_in_time:
                    check_in_str = record.check_in_time.astimezone(vietnam_tz).strftime('%H:%M')
                if record.check_out_time:
                    check_out_str = record.check_out_time.astimezone(vietnam_tz).strftime('%H:%M')
                
                status_display = record.get_status_display()
                
                row_data = [
                    dept.name,
                    emp.employee_id,
                    emp.get_full_name(),
                    record.date.strftime('%d/%m/%Y'),
                    check_in_str,
                    check_out_str,
                    status_display
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws2.cell(row=detail_row, column=col, value=value)
                    cell.border = thin_border
                    if col in [4, 5, 6, 7]:
                        cell.alignment = Alignment(horizontal="center")
                    # Apply status color to status column
                    if col == 7:
                        status_fill = status_fills.get(record.status)
                        if status_fill:
                            cell.fill = status_fill
                
                detail_row += 1
    
    # Adjust column widths for sheet 2
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 10
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['G'].width = 15
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Create response
    filename = f"thong_ke_cham_cong_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@csrf_exempt
def department_employees_attendance_api(request, department_name):
    """API to get detailed employee attendance for a specific department"""
    if request.method != 'GET':
        return error_response('Method not allowed', 405)
    
    from datetime import datetime
    
    vietnam_now = get_vietnam_now()
    vietnam_tz = vietnam_now.tzinfo
    
    # Get filter parameters
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Default to current month
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = vietnam_now.date().replace(day=1)
    else:
        start_date = vietnam_now.date().replace(day=1)
    
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            end_date = vietnam_now.date()
    else:
        end_date = vietnam_now.date()
    
    # URL decode department name
    from urllib.parse import unquote
    department_name = unquote(department_name)
    
    # Get employees in this department
    employees = Employee.objects.filter(
        department=department_name,
        work_status='WORKING'
    ).select_related('user')
    
    employee_data = []
    for emp in employees:
        # Get attendance records for this employee within date range
        records = AttendanceRecord.objects.filter(
            employee=emp,
            date__gte=start_date,
            date__lte=end_date
        ).order_by('-date')
        
        # Count stats
        on_time = records.filter(status='ON_TIME').count()
        late = records.filter(status='LATE').count()
        early = records.filter(status='EARLY').count()
        total_records = records.count()
        on_time_rate = (on_time / total_records * 100) if total_records > 0 else 0
        
        # Get attendance history
        history = []
        for record in records[:30]:  # Limit to 30 records
            check_in_vn = None
            check_out_vn = None
            if record.check_in_time:
                check_in_vn = record.check_in_time.astimezone(vietnam_tz).strftime('%H:%M')
            if record.check_out_time:
                check_out_vn = record.check_out_time.astimezone(vietnam_tz).strftime('%H:%M')
            
            history.append({
                'date': record.date.isoformat(),
                'date_display': record.date.strftime('%d/%m/%Y'),
                'check_in': check_in_vn,
                'check_out': check_out_vn,
                'status': record.status,
                'status_display': record.get_status_display(),
            })
        
        employee_data.append({
            'employee_id': emp.employee_id,
            'full_name': emp.get_full_name(),
            'position': emp.position,
            'total_records': total_records,
            'on_time': on_time,
            'late': late,
            'early': early,
            'on_time_rate': round(on_time_rate, 1),
            'attendance_history': history,
        })
    
    # Sort by late count descending (show employees with most late first)
    employee_data.sort(key=lambda x: x['late'], reverse=True)
    
    return json_response({
        'success': True,
        'department': department_name,
        'employees': employee_data,
        'filters': {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
        },
    })


